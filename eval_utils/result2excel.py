from openpyxl import Workbook
import os
import numpy as np

wb = Workbook()

sheet1 = wb.active
sheet2 = wb.create_sheet()
sheet3 = wb.create_sheet()
sheet4 = wb.create_sheet()
sheet5 = wb.create_sheet()
sheet6 = wb.create_sheet()
file_name = 'results_PhDThesis19.xlsx'


r_title = ['Noisy', 'IMCRA', 'DRNN', 'U-Net', 'SEGAN', \
           'SNMF', 'SNMF+ONDL', 'SNMF+ONDL+SPSC',  \
           'TAU-Net', 'TAU-Net+SPSC']
c_title = ['0dB', '5dB', '10dB', '15dB', 'Avg.']

TABLE_GAP = len(r_title) + 2
sheet1.title = 'obj_Thesis_SDR'
sheet2.title = 'obj_Thesis_SIR'
sheet3.title = 'obj_Thesis_SAR'
sheet4.title = 'obj_Thesis_PESQ'
sheet5.title = 'obj_Thesis_STOI'
sheet6.title = 'obj_Thesis_SSNR'
#Mu Table
for r_idx in range(1,len(r_title)+1):
    sheet1.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]
    sheet2.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]
    sheet3.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]
    sheet4.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]
    sheet5.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]
    sheet6.cell(row=r_idx+1, column=1).value = r_title[r_idx-1]

for c_idx in range(1,len(c_title)+1):
    sheet1.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]
    sheet2.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]
    sheet3.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]
    sheet4.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]
    sheet5.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]
    sheet6.cell(row=1, column=c_idx+1).value = c_title[c_idx-1]

#Var Table
for r_idx in range(1,len(r_title)+1):
    sheet1.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]
    sheet2.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]
    sheet3.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]
    sheet4.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]
    sheet5.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]
    sheet6.cell(row=r_idx+1 + TABLE_GAP, column=1).value = r_title[r_idx-1]

for c_idx in range(1,len(c_title)+1):
    sheet1.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]    
    sheet2.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]   
    sheet3.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]   
    sheet4.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]   
    sheet5.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]     
    sheet6.cell(row=1 + TABLE_GAP, column=c_idx+1).value = c_title[c_idx-1]     


RESULT_PATH = 'results/Thesis_TIMIT'
for (path, dir, files) in os.walk(RESULT_PATH):
    for file in files:
        [r_name, c_name, exc] = file.split('_')
        for r in range(0,len(r_title)):
            if r_title[r] == r_name:
                r_idx = r+2
        for c in range(0,len(c_title)):
            if c_title[c].split('dB')[0] == c_name:
                c_idx = c+2

        f = open(path + '/' + file, 'rt')
        val = f.readlines() #[SDR, SIR, SAR, PESQ]
        [mu, var] = val[0].split(', ') 
        sheet1.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet1.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        [mu, var] = val[1].split(', ') 
        sheet2.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet2.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        [mu, var] = val[2].split(', ') 
        sheet3.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet3.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        [mu, var] = val[3].split(', ') 
        sheet4.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet4.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        [mu, var] = val[4].split(', ') 
        sheet5.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet5.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        [mu, var] = val[5].split(', ') 
        sheet6.cell(row=r_idx, column=c_idx).value = float(mu)
        sheet6.cell(row=r_idx+TABLE_GAP, column=c_idx).value = np.sqrt(float(var) * 0.5)
        

wb.save(filename=file_name)


